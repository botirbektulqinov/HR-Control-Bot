"""Enterprise ko'rinishdagi Excel davomat hisoboti (openpyxl)."""

from __future__ import annotations

import datetime as dt
import io
from collections import Counter
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Attendance, Employee
from app.i18n import status_label

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
STATUS_FILL = {
    "on_time": "C6EFCE",
    "late": "FFEB9C",
    "very_late": "FFC7CE",
    "weekend": "D9D9D9",
}
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")

HEADERS = [
    "№", "F.I.Sh", "Bo'lim", "Filial", "Sana", "Keldi", "Ketdi",
    "Ishlangan (soat)", "Kechikish (min)", "Erta ketish (min)",
    "Overtime (min)", "Holat", "Masofa (m)",
]


def _hm(when: dt.datetime | None, tz: str) -> str:
    return when.astimezone(ZoneInfo(tz)).strftime("%H:%M") if when else ""


async def _fetch(session: AsyncSession, d_from: dt.date, d_to: dt.date):
    q = (
        select(Attendance)
        .join(Employee)
        .where(Attendance.work_date >= d_from, Attendance.work_date <= d_to)
        .order_by(Employee.full_name, Attendance.work_date)
    )
    return (await session.scalars(q)).all()


async def build_report(
    session: AsyncSession, d_from: dt.date, d_to: dt.date, lang: str = "uz"
) -> tuple[str, io.BytesIO] | None:
    rows = await _fetch(session, d_from, d_to)
    if not rows:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    ws["A1"] = f"Davomat hisoboti:  {d_from} — {d_to}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.append([])  # 2-qator bo'sh
    ws.append(HEADERS)  # 3-qator sarlavha
    hr = 3
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER

    widths = [len(h) for h in HEADERS]
    counter: Counter[str] = Counter()
    for i, r in enumerate(rows, start=1):
        emp = r.employee
        tz = emp.branch.timezone if emp.branch else "Asia/Tashkent"
        counter[r.status] += 1
        values = [
            i, emp.full_name, emp.department or "", emp.branch.name if emp.branch else "",
            str(r.work_date), _hm(r.check_in_at, tz), _hm(r.check_out_at, tz),
            round(r.worked_minutes / 60, 2), r.late_minutes, r.early_leave_minutes,
            r.overtime_minutes, status_label(lang, r.status),
            round(r.check_in_distance_m) if r.check_in_distance_m is not None else "",
        ]
        ws.append(values)
        row_idx = hr + i
        fill = STATUS_FILL.get(r.status)
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            if c in (1, 5, 6, 7, 8, 9, 10, 11, 13):
                cell.alignment = CENTER
            if c == 12 and fill:  # Holat ustuni rangli
                cell.fill = PatternFill("solid", fgColor=fill)
            widths[c - 1] = max(widths[c - 1], len(str(v)))

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(w + 2, 40)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(HEADERS))}{hr}"

    # Xulosa varag'i
    s = wb.create_sheet("Xulosa")
    s["A1"] = "Xulosa"
    s["A1"].font = Font(bold=True, size=14, color="1F4E78")
    s.append([])
    s.append(["Ko'rsatkich", "Qiymat"])
    for c in (1, 2):
        s.cell(row=3, column=c).fill = HEADER_FILL
        s.cell(row=3, column=c).font = HEADER_FONT
    total_late = sum(r.late_minutes for r in rows)
    summary = [
        ("Jami yozuvlar", len(rows)),
        ("O'z vaqtida", counter.get("on_time", 0)),
        ("Kechikkan", counter.get("late", 0) + counter.get("very_late", 0)),
        ("Jami kechikish (min)", total_late),
    ]
    for name, val in summary:
        s.append([name, val])
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return f"davomat_{d_from}_{d_to}.xlsx", buf
